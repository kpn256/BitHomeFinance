
from openpyxl import load_workbook
import requests
import pandas as pd




def infoInput(coin,date,sats,fiat):


    wb = load_workbook('datos2.xlsx')
    ws = wb.active
    ws.append([date,sats,fiat])
    wb.save('datos2.xlsx')
       
    dca = load_workbook("datos2.xlsx")
    ws = dca.active
    maxrow = dca.active.max_row
    minrow = dca.active.min_row
    ws['D2'] = f'=SUM(b{minrow + 1}:b{maxrow})'
    ws['E2'] = f'=SUM(c{minrow + 1}:c{maxrow})'
    dca.save("datos2.xlsx")

    df_dca = pd.read_excel("datos2.xlsx")
    totalsats = df_dca['SATS'].sum()
    totalfiat = df_dca['FIAT'].sum()
    sats = totalsats
    sats = sats / 100000000
    url = (f'https://api.yadio.io/convert/{sats}/BTC/{coin}')
    response = requests.get(url)
    Value = response.json()
    result = Value["result"]
    actualPrice = result

    return f"total sats in DCA: {totalsats}\n total {coin} spend: {totalfiat}\n {totalsats}sats = {actualPrice} {coin}"
    

def total(coin):

    df_dca = pd.read_excel("datos2.xlsx")
    totalsats = df_dca['SATS'].sum()
    totalfiat = df_dca['FIAT'].sum()
    sats = totalsats
    sats = sats / 100000000
    url = (f'https://api.yadio.io/convert/{sats}/BTC/{coin}')
    response = requests.get(url)
    Value = response.json()
    result = Value["result"]
    actualPrice = result
    print(f"total sats in DCA: {totalsats}\n"
          f"total {coin} spend: {totalfiat}\n"
          f"{totalsats} sat = {actualPrice} {coin}")




from openpyxl import load_workbook
import pandas as pd
import os


def pfLogic(election, date, amount, note, ):

    

    wb = load_workbook("datos1.xlsx")
    spends = wb["spends"]
    incomes = wb["incomes"]

    if election == "s":
        spends.append([date,amount,note])
        

    elif election == "i":
        incomes.append([date,amount,note])
        
    maxrowS = wb["spends"].max_row
    maxrowI = wb["incomes"].max_row
    spends['D2'] = f'=SUM(b2:b{maxrowS})'
    incomes['D2'] = f'=SUM(b2:b{maxrowI})'
    wb.save("datos1.xlsx")
    

def total():
    datos = pd.read_excel("datos1.xlsx",sheet_name="spends")
    totalSpends = datos['AMOUNT'].sum()
    datos = pd.read_excel("datos1.xlsx",sheet_name="incomes")
    totalIncomes = datos['AMOUNT'].sum()

    return f"total spends :{totalSpends}\n total income: {totalIncomes}\n"

